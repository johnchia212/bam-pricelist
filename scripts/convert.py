import openpyxl, json, re, datetime

SRC = "/sessions/exciting-youthful-wright/mnt/uploads/Price_2026_Export MASTER V260105.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

HEADERS = ["series","model","sap_pn","capacity","dimensions","range","weight_kg","et_mm","hcg_mm","mounting_class","price_rmb","updated","remarks"]

def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.replace("\xa0", " ").strip()
        v = v.strip("'")
        return v if v != "" else None
    return v

categories = []
total_rows = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        cleaned = [clean(v) for v in row[:13]]
        if all(v is None for v in cleaned):
            continue
        rec = dict(zip(HEADERS, cleaned))
        # normalize price to number
        if isinstance(rec.get("price_rmb"), str):
            try:
                rec["price_rmb"] = float(re.sub(r"[^\d.]", "", rec["price_rmb"]))
            except Exception:
                pass
        rec["category"] = sheet_name
        rows.append(rec)
    total_rows += len(rows)
    categories.append({"category": sheet_name, "count": len(rows), "items": rows})

out = {
    "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
    "source_file": "Price_2026_Export MASTER V260105.xlsx",
    "total_items": total_rows,
    "categories": categories
}

with open("/sessions/exciting-youthful-wright/mnt/outputs/bam-pricelist/data/pricelist.json", "w") as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print("Total items:", total_rows)
print("Categories:", [(c["category"], c["count"]) for c in categories])
